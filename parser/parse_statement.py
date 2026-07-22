#!/usr/bin/env python3
"""
parse_statement.py
Generic bank-statement PDF -> XLSX converter.

Works across banks by auto-detecting the table header row on each page
(Date / Particulars / Withdrawal / Deposit / Balance, under whatever
label wording a given bank uses) instead of assuming fixed column
positions. Groups wrapped, multi-line narrations under the correct
transaction using the date column as an anchor, then reconciles every
row's balance against (previous balance - withdrawal + deposit) so you
know, mechanically, whether the extraction is trustworthy -- not just
by eyeballing it.

Usage:
    python3 parse_statement.py --input statement.pdf --output out.xlsx \
        --status-file status.json [--job-id 123]

Writes progress + final results to --status-file as JSON so a PHP
front end can poll it without holding an HTTP request open (avoids
PHP execution-time-limit / gateway timeouts on long statements).

Exit code 0 = parsed (check status.json "status" field: "completed"
or "completed_with_warnings"). Exit code 1 = failed; see "error".
"""

import argparse
import json
import re
import sys
import traceback
from datetime import date, datetime

try:
    import pdfplumber
except ImportError:
    print("Missing dependency: pip install pdfplumber", file=sys.stderr)
    raise

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Header keyword synonyms -- this is what makes the parser bank-agnostic.
# Add more synonyms here as you feed it statements from new banks.
# --------------------------------------------------------------------------
HEADER_SYNONYMS = {
    "date":        ["date", "txn date", "transaction date", "value date"],
    "particulars": ["particulars", "narration", "description", "details",
                     "transaction details", "remarks", "transaction remarks"],
    "chqno":       ["chq.no", "chq no", "cheque no", "chq/ref.no", "ref no",
                     "reference no", "chq./ref.no", "instrument no"],
    "withdrawal":  ["withdrawals", "withdrawal", "debit", "dr amount",
                     "withdrawal amt", "withdrawal amount"],
    "deposit":     ["deposits", "deposit", "credit", "cr amount",
                     "deposit amt", "deposit amount"],
    "amount":      ["amount", "transaction amount"],   # single-column style
    "drcr":        ["dr/cr", "cr/dr", "type", "dc"],    # single-column style
    "balance":     ["balance", "closing balance", "running balance",
                     "available balance"],
}

# Cosmetic only -- used to label the output, never to pick a column layout.
BANK_NAME_HINTS = [
    "SOUTH INDIAN BANK", "STATE BANK OF INDIA", "HDFC BANK", "ICICI BANK",
    "AXIS BANK", "KOTAK MAHINDRA", "PUNJAB NATIONAL BANK", "BANK OF BARODA",
    "CANARA BANK", "UNION BANK OF INDIA", "IDBI BANK", "YES BANK",
    "INDUSIND BANK", "FEDERAL BANK", "KARNATAKA BANK", "CITY UNION BANK",
    "BANDHAN BANK", "IDFC FIRST BANK",
]

NOISE_LINE_PATTERNS = [
    r"^page\s*total", r"^grand\s*total", r"^total\s*:", r"^b/f\b",
    r"^c/f\b", r"brought\s*forward", r"carried\s*forward",
    r"^opening\s*balance", r"^closing\s*balance",
    r"statement\s*of\s*account", r"generated\s*statement",
    r"^visit\s*us", r"customer\s*care", r"^page\s+\d+\s+of\s+\d+",
    r"this\s+is\s+a\s+system\s+generated", r"^date/time\s*:",
]
NOISE_RE = re.compile("|".join(NOISE_LINE_PATTERNS), re.IGNORECASE)

# Once any of these appear, everything on the page below them is statement
# boilerplate (footnotes, legends, disclaimers) -- stop collecting entirely
# rather than letting it get glued onto the last real transaction's narration.
STOP_MARKER_RE = re.compile(
    r"grand\s*total|date format used|abbreviations used|important message|"
    r"^disclaimer|deposit insurance and credit guarantee",
    re.IGNORECASE,
)

# Bare transaction-type / channel codes that sometimes sit in their own
# column with no dedicated header synonym, and end up folded into the
# narration column as noise. Only stripped when they appear as an
# isolated whole word, never as part of a longer word.
NARRATION_NOISE_WORDS = {
    "TFR", "CASH", "CLG", "MB", "FT", "RTG", "CHRG", "NEFT", "IMPS", "SBINT",
}

DATE_PATTERNS = [
    re.compile(r"^\d{2}-\d{2}-\d{2,4}$"),
    re.compile(r"^\d{2}/\d{2}/\d{2,4}$"),
    re.compile(r"^\d{1,2}-[A-Za-z]{3}-\d{2,4}$"),
    re.compile(r"^\d{1,2}\s[A-Za-z]{3}\s\d{2,4}$"),
]
AMT_RE = re.compile(r"^\(?-?[\d,]+(\.\d{1,2})?\)?(Cr|Dr|CR|DR)?$")


def log_status(status_file, **kwargs):
    """Write current progress to the status file for the PHP layer to poll."""
    payload = {}
    try:
        with open(status_file) as f:
            payload = json.load(f)
    except Exception:
        pass
    payload.update(kwargs)
    payload["updated_at"] = datetime.utcnow().isoformat() + "Z"
    tmp = status_file + ".tmp"
    with open(tmp, "w") as f:
        json.dump(payload, f, indent=2, default=str)
    import os
    os.replace(tmp, status_file)


def is_date_token(text):
    return any(p.match(text) for p in DATE_PATTERNS)


def is_amount_token(text):
    return bool(AMT_RE.match(text))


def parse_amount(s):
    if not s:
        return None
    s = s.strip().replace("Cr", "").replace("Dr", "").replace("CR", "").replace("DR", "")
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "")
    if not s:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def parse_date_token(s, default_century_cutoff=50):
    s = s.strip()
    fmts = ["%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y",
            "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %b %y"]
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def cluster_lines(words, gap=3.0):
    """Group words into visual lines by vertical position (top), tolerant
    of the small sub-pixel offsets PDF renderers introduce between words
    that are meant to sit on the same printed line."""
    words = sorted(words, key=lambda w: w["top"])
    lines, cur, cur_top = [], [], None
    for w in words:
        if cur and w["top"] - cur_top > gap:
            lines.append(cur)
            cur = []
        cur.append(w)
        cur_top = w["top"]
    if cur:
        lines.append(cur)
    return lines


def detect_header(lines):
    """Scan clustered lines for a row containing several of the expected
    column labels. Returns dict: field -> x0 (left edge) of that column,
    or None if no header row is found on this page."""
    best = None
    best_score = 0
    for ln in lines:
        ln_sorted = sorted(ln, key=lambda w: w["x0"])
        matches = {}
        for w in ln_sorted:
            wl = w["text"].strip(":.").lower()
            for field, syns in HEADER_SYNONYMS.items():
                for syn in syns:
                    if wl == syn or (len(syn.split()) > 1 and syn in
                                      " ".join(t["text"].lower() for t in ln_sorted)):
                        if field not in matches:
                            matches[field] = w["x0"]
        score = len(matches)
        if score > best_score and "date" in matches and "particulars" in matches:
            best_score = score
            best = matches
    return best if best_score >= 3 else None


def make_midpoint_boundaries(boundaries):
    """boundaries: sorted list of (x0_start, field) pairs using each
    column's LABEL position. Amount columns are right-aligned, so a
    given cell's actual text can start well to the left of its header
    label (e.g. '58,646.41Cr' starts left of where 'BALANCE' is
    printed). Splitting on the midpoint between consecutive label
    positions -- rather than the label position itself -- classifies
    those right-aligned values correctly."""
    fields = [f for _, f in boundaries]
    starts = [x0 for x0, _ in boundaries]
    cuts = [-1e9]
    for i in range(1, len(starts)):
        cuts.append((starts[i - 1] + starts[i]) / 2)
    return list(zip(cuts, fields))


def classify_x0(x0, midpoint_boundaries):
    """midpoint_boundaries: sorted list of (cut_start, field) pairs built
    by make_midpoint_boundaries(). Returns the field whose cut_start is
    the greatest value <= x0."""
    field = midpoint_boundaries[0][1]
    for start, f in midpoint_boundaries:
        if x0 >= start:
            field = f
        else:
            break
    return field


def extract_transactions(pdf_path, status_file):
    transactions = []
    bank_name = None
    all_page_text_sample = ""
    current_header = None

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        for pnum, page in enumerate(pdf.pages, 1):
            if pnum % 5 == 0 or pnum == 1:
                log_status(status_file, status="processing",
                           message=f"Reading page {pnum} of {total_pages}",
                           progress=round(pnum / total_pages * 90, 1))

            words = page.extract_words()
            if not words:
                continue  # blank / image-only page

            if pnum == 1:
                header_zone_words = [w["text"] for w in words if w["top"] < 150]
                all_page_text_sample = " ".join(header_zone_words)

            lines = cluster_lines(words)
            header = detect_header(lines)
            if header:
                current_header = header
            if not current_header:
                continue  # haven't found a header yet; nothing parseable

            # Build column boundaries sorted by x0, then convert to
            # midpoint cut-points (see make_midpoint_boundaries docstring).
            boundaries = sorted(current_header.items(), key=lambda kv: kv[1])
            boundaries = [(x0, f) for f, x0 in boundaries]
            boundaries = make_midpoint_boundaries(boundaries)

            page_height = page.height
            content_words = []
            for ln in lines:
                ln_sorted = sorted(ln, key=lambda w: w["x0"])
                text = " ".join(w["text"] for w in ln_sorted)
                if STOP_MARKER_RE.search(text):
                    break  # rest of page is boilerplate/footnotes -- stop entirely
                if NOISE_RE.search(text):
                    continue
                # A line that is nothing but a short bare number sitting in
                # the bottom margin is a folio/page number (e.g. just "3"
                # printed bottom-right), not a transaction value -- without
                # this it gets misclassified as that page's last Balance.
                if (re.match(r"^\d{1,4}$", text.strip())
                        and ln_sorted[0]["top"] > page_height - 70):
                    continue
                content_words.extend(ln_sorted)

            # Determine date column's x-range (left of particulars boundary)
            date_x0 = current_header.get("date", 0)
            particulars_x0 = current_header.get("particulars", 999999)

            anchors = []
            for w in content_words:
                if w["x0"] < particulars_x0 - 2 and is_date_token(w["text"]):
                    anchors.append(w["top"])
            anchors = sorted(set(anchors))
            if not anchors:
                continue

            for i, a_top in enumerate(anchors):
                lo = a_top - 3
                hi = anchors[i + 1] - 3 if i + 1 < len(anchors) else 1e9
                tx_words = [w for w in content_words if lo <= w["top"] < hi]

                date_w = [w for w in tx_words
                          if w["x0"] < particulars_x0 - 2 and is_date_token(w["text"])]
                date_val = date_w[0]["text"] if date_w else None

                buckets = {k: [] for k in
                           ["particulars", "chqno", "withdrawal", "deposit", "amount", "drcr", "balance"]}

                for w in sorted(tx_words, key=lambda w: (w["top"], w["x0"])):
                    if w in date_w:
                        continue
                    field = classify_x0(w["x0"], boundaries)
                    if field in buckets:
                        buckets[field].append(w["text"])
                    elif field == "date":
                        # stray token in date column that isn't the anchor itself
                        buckets["particulars"].append(w["text"])

                # Resolve withdrawal/deposit either from dedicated columns
                # or from a single amount + Dr/Cr indicator column.
                withdrawal = None
                deposit = None
                if buckets["withdrawal"] or buckets["deposit"]:
                    amt_w = [t for t in buckets["withdrawal"] if is_amount_token(t)]
                    amt_d = [t for t in buckets["deposit"] if is_amount_token(t)]
                    withdrawal = amt_w[-1] if amt_w else None
                    deposit = amt_d[-1] if amt_d else None
                elif buckets["amount"]:
                    amt_tokens = [t for t in buckets["amount"] if is_amount_token(t)]
                    amt = amt_tokens[-1] if amt_tokens else None
                    drcr_tokens = [t.upper() for t in buckets["drcr"]]
                    is_dr = any("DR" in t or t == "D" for t in drcr_tokens)
                    is_cr = any("CR" in t or t == "C" for t in drcr_tokens)
                    if amt and "CR" in amt.upper():
                        is_cr = True
                    if amt and "DR" in amt.upper():
                        is_dr = True
                    if amt:
                        if is_dr and not is_cr:
                            withdrawal = amt
                        else:
                            deposit = amt

                bal_tokens = [t for t in buckets["balance"] if is_amount_token(t)]
                balance = bal_tokens[-1] if bal_tokens else None

                narration_words = [t for t in buckets["particulars"] if t not in NARRATION_NOISE_WORDS]
                narration = " ".join(narration_words).strip()
                chqno = " ".join(buckets["chqno"]).strip()

                if not date_val and not narration and not balance:
                    continue  # nothing usable; skip stray fragment

                transactions.append({
                    "page": pnum,
                    "date_raw": date_val,
                    "particulars": narration,
                    "chqno": chqno,
                    "withdrawal": withdrawal,
                    "deposit": deposit,
                    "balance": balance,
                })

    for hint in BANK_NAME_HINTS:
        if hint in all_page_text_sample.upper():
            bank_name = hint.title()
            break

    return transactions, bank_name


def reconcile(transactions):
    """Walk the parsed rows and verify prev_balance - withdrawal + deposit
    == this row's balance. Returns (mismatch_count, mismatch_details)."""
    prev_bal = None
    mismatches = []
    for i, t in enumerate(transactions):
        w = parse_amount(t["withdrawal"]) or 0.0
        d = parse_amount(t["deposit"]) or 0.0
        bal = parse_amount(t["balance"])
        if bal is None:
            mismatches.append({"index": i, "reason": "missing balance", "row": t})
            continue
        if prev_bal is not None:
            expected = prev_bal - w + d
            if abs(expected - bal) > 0.02:
                mismatches.append({
                    "index": i, "reason": "balance mismatch",
                    "expected": round(expected, 2), "actual": round(bal, 2), "row": t,
                })
        prev_bal = bal
    return mismatches


def write_xlsx(transactions, output_path, bank_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    FONT_NAME = "Arial"
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name=FONT_NAME, size=10)
    bold_font = Font(name=FONT_NAME, size=10, bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    warn_fill = PatternFill(start_color="FDE9E9", end_color="FDE9E9", fill_type="solid")

    ws["A1"] = f"BANK STATEMENT — {bank_name or 'AUTO-DETECTED'}"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
    ws["A2"] = "Converted automatically. Rows highlighted in red failed balance reconciliation — verify against the source PDF."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="595959")

    header_row = 4
    headers = ["S.No", "Date", "Particulars", "Cheque No.", "Withdrawal (Dr)", "Deposit (Cr)", "Balance"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    prev_bal = None
    r = header_row + 1
    for idx, t in enumerate(transactions, start=1):
        w = parse_amount(t["withdrawal"])
        d = parse_amount(t["deposit"])
        b = parse_amount(t["balance"])
        dt = parse_date_token(t["date_raw"]) if t["date_raw"] else None

        mismatch = False
        if b is not None and prev_bal is not None:
            expected = prev_bal - (w or 0.0) + (d or 0.0)
            if abs(expected - b) > 0.02:
                mismatch = True
        if b is not None:
            prev_bal = b

        ws.cell(row=r, column=1, value=idx)
        c_date = ws.cell(row=r, column=2, value=dt if dt else t["date_raw"])
        if dt:
            c_date.number_format = "dd-mm-yyyy"
        ws.cell(row=r, column=3, value=t["particulars"])
        ws.cell(row=r, column=4, value=t["chqno"] or None)
        ws.cell(row=r, column=5, value=w)
        ws.cell(row=r, column=6, value=d)
        ws.cell(row=r, column=7, value=b)
        for col in (5, 6, 7):
            ws.cell(row=r, column=col).number_format = "#,##0.00"

        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.font = normal_font
            cell.border = border
            if mismatch:
                cell.fill = warn_fill
            elif idx % 2 == 0:
                cell.fill = alt_fill
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        r += 1

    last_data_row = r - 1
    ws.cell(row=r, column=3, value="TOTAL").font = bold_font
    ws.cell(row=r, column=5, value=f"=SUM(E{header_row+1}:E{last_data_row})")
    ws.cell(row=r, column=6, value=f"=SUM(F{header_row+1}:F{last_data_row})")
    ws.cell(row=r, column=7, value=f"=G{last_data_row}")
    for col in (3, 5, 6, 7):
        cell = ws.cell(row=r, column=col)
        cell.font = bold_font
        cell.border = border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        if col in (5, 6, 7):
            cell.number_format = "#,##0.00"

    widths = {1: 7, 2: 12, 3: 65, 4: 14, 5: 16, 6: 16, 7: 16}
    for col, wd in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = wd
    ws.freeze_panes = f"A{header_row+1}"
    ws.row_dimensions[header_row].height = 28

    wb.save(output_path)
    return last_data_row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--analysis-output", required=False, default=None,
                     help="Path to write the self-contained HTML audit/ITR analysis report.")
    ap.add_argument("--status-file", required=True)
    ap.add_argument("--job-id", default="")
    args = ap.parse_args()

    log_status(args.status_file, status="processing", progress=0,
               message="Starting", job_id=args.job_id)

    try:
        transactions, bank_name = extract_transactions(args.input, args.status_file)

        if not transactions:
            log_status(args.status_file, status="failed",
                       error="No transactions could be extracted. The PDF may be a "
                             "scanned image (no text layer) or use an unrecognised "
                             "table layout.", progress=100)
            sys.exit(1)

        log_status(args.status_file, status="processing", progress=93,
                   message="Reconciling balances")
        mismatches = reconcile(transactions)

        log_status(args.status_file, status="processing", progress=96,
                   message="Writing Excel file")
        write_xlsx(transactions, args.output, bank_name)

        analysis_file = None
        flags_count = None
        if args.analysis_output:
            log_status(args.status_file, status="processing", progress=98,
                       message="Building audit / ITR analysis report")
            import os as _os
            from analyze_statement import build_analysis, render_html

            analysis_rows = []
            for t in transactions:
                analysis_rows.append({
                    "date": parse_date_token(t["date_raw"]) if t["date_raw"] else None,
                    "particulars": t["particulars"],
                    "withdrawal": parse_amount(t["withdrawal"]),
                    "deposit": parse_amount(t["deposit"]),
                    "balance": parse_amount(t["balance"]),
                })
            analysis = build_analysis(analysis_rows, bank_name)
            report_html = render_html(analysis, {
                "source_file": _os.path.basename(args.input),
                "generated_at": datetime.now().strftime("%d-%b-%Y %H:%M"),
            })
            with open(args.analysis_output, "w", encoding="utf-8") as f:
                f.write(report_html)
            analysis_file = args.analysis_output
            flags_count = len(analysis["flags"])

        total_w = sum(parse_amount(t["withdrawal"]) or 0.0 for t in transactions)
        total_d = sum(parse_amount(t["deposit"]) or 0.0 for t in transactions)
        closing_balance = None
        for t in reversed(transactions):
            v = parse_amount(t["balance"])
            if v is not None:
                closing_balance = v
                break
        opening_balance = None
        for t in transactions:
            v = parse_amount(t["balance"])
            if v is not None:
                w = parse_amount(t["withdrawal"]) or 0.0
                d = parse_amount(t["deposit"]) or 0.0
                opening_balance = v + w - d
                break

        status = "completed" if not mismatches else "completed_with_warnings"
        log_status(
            args.status_file,
            status=status,
            progress=100,
            message="Done",
            bank_name=bank_name,
            transactions_count=len(transactions),
            total_withdrawals=round(total_w, 2),
            total_deposits=round(total_d, 2),
            opening_balance=round(opening_balance, 2) if opening_balance is not None else None,
            closing_balance=round(closing_balance, 2) if closing_balance is not None else None,
            balance_mismatches=len(mismatches),
            mismatch_sample=[
                {"index": m["index"], "reason": m["reason"],
                 "expected": m.get("expected"), "actual": m.get("actual")}
                for m in mismatches[:20]
            ],
            output_file=args.output,
            analysis_file=analysis_file,
            audit_flags_count=flags_count,
        )
        sys.exit(0)

    except Exception as e:
        log_status(args.status_file, status="failed",
                   error=str(e), traceback=traceback.format_exc(), progress=100)
        sys.exit(1)


if __name__ == "__main__":
    main()
